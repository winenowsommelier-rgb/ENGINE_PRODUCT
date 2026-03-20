#!/usr/bin/env python3
"""Extract all taxonomy sheets from Excel to JSON files."""

import openpyxl
import json
from pathlib import Path

def extract_taxonomy():
    file_path = Path("Upload/full_global_taxonomy_pro (1).xlsx")
    wb = openpyxl.load_workbook(file_path)
    output_dir = Path("data/taxonomy")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Extract each taxonomy sheet with proper headers
    taxonomy_sheets = {
        'countries': ['id', 'name', 'iso'],
        'regions': ['id', 'country_id', 'name'],
        'subregions': ['id', 'region_id', 'name', 'subregion_type'],
        'Origin': ['id', 'subregion_id', 'name', 'origin_type'],
        'classification_master': ['classification_id', 'classification', 'classification_slug', 
                                  'classification_group', 'category_scope', 'priority', 'description', 'is_active'],
        'ingredient_master': ['ingredient_id', 'ingredient', 'ingredient_slug', 
                              'ingredient_group', 'category_scope', 'is_primary_default', 'synonyms', 'description', 'is_active'],
        'flavor_note_master': ['note_id', 'note', 'note_slug', 'note_family', 'is_active'],
        'category_render_config': ['category', 'show_gauges', 'show_wheel', 'show_matrix', 
                                   'primary_gauge_keys', 'matrix_x_left_label', 'matrix_x_right_label', 
                                   'matrix_y_bottom_label', 'matrix_y_top_label', 'recommended_blocks', 'is_active'],
        'expert_sources': ['id', 'name', 'Score'],
    }

    for sheet_name, headers in taxonomy_sheets.items():
        ws = wb[sheet_name]
        rows = []
        
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row_idx, col_idx).value
                row_data[header] = cell_value
            
            if any(v is not None for v in row_data.values()):
                rows.append(row_data)
        
        output_file = output_dir / f"{sheet_name.lower().replace(' ', '_')}.json"
        with open(output_file, 'w') as f:
            json.dump({
                'sheet': sheet_name,
                'count': len(rows),
                'data': rows
            }, f, indent=2, default=str)
        
        print(f"✓ {sheet_name}: {len(rows)} rows → {output_file.name}")

if __name__ == '__main__':
    extract_taxonomy()
    print("\nAll taxonomy sheets extracted to data/taxonomy/")
