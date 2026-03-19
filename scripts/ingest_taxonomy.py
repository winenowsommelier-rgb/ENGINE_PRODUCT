#!/usr/bin/env python3
"""Convert taxonomy workbooks into JSON outputs.

This script loads the provided Excel workbook, exports each tab as a JSON file, and
provides a small summary for the "Magento item data" sheet (which contains product SKU
records).

Usage:
  python scripts/ingest_taxonomy.py \
    --workbook Upload/full_global_taxonomy_pro\ (1).xlsx \
    --out data/taxonomy

The script is intentionally lightweight and does not attempt to infer or normalize
values beyond basic dtype conversion (numbers, strings, booleans).
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl


def slugify_name(name: str) -> str:
    # Keep it simple: lowercase, replace spaces and slashes, and remove parentheses.
    return (
        name
        .strip()
        .lower()
        .replace(' ', '_')
        .replace('/', '_')
        .replace('(', '')
        .replace(')', '')
        .replace('-', '_')
    )


def normalize_header(header: Any, index: int) -> str:
    if header is None:
        return f'column_{index}'
    header_str = str(header).strip()
    if not header_str:
        return f'column_{index}'
    # Make safe for keys; allow duplicates by appending index when needed.
    return header_str


def sheet_to_records(sheet: openpyxl.worksheet.worksheet.Worksheet) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Convert an openpyxl worksheet into a list of records (dicts)."""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers: List[str] = []
    seen: Dict[str, int] = {}
    for idx, raw in enumerate(rows[0]):
        hdr = normalize_header(raw, idx)
        key = hdr
        if key in seen:
            seen[key] += 1
            key = f"{key}_{seen[key]}"
        else:
            seen[key] = 0
        headers.append(key)

    records: List[Dict[str, Any]] = []
    for row in rows[1:]:
        record: Dict[str, Any] = {}
        for key, value in zip(headers, row):
            record[key] = value
        records.append(record)
    return headers, records


def dump_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet, out_dir: Path) -> None:
    headers, records = sheet_to_records(sheet)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{slugify_name(sheet.title)}.json"
    with out_path.open('w', encoding='utf-8') as f:
        json.dump({'sheet': sheet.title, 'headers': headers, 'rowCount': len(records), 'rows': records}, f, indent=2, default=str)
    print(f'Wrote {out_path} ({len(records)} rows, {len(headers)} columns)')


def summarize_magento_items(records: List[Dict[str, Any]]) -> None:
    print('\nSummary: Magento item data')
    print('Total rows:', len(records))
    if not records:
        return

    keys = list(records[0].keys())
    print('Columns:', ', '.join(keys))
    sample = records[:3]
    print('\nSample rows:')
    for r in sample:
        # Show a small subset of fields for readability
        subset = {k: r.get(k) for k in keys if k in ('sku', 'name', 'price', 'cost', 'country', 'region_wine', 'grape_variety', 'wine_type')}
        print('  ', subset)


def main() -> None:
    parser = argparse.ArgumentParser(description='Convert taxonomy workbook into JSON exports.')
    parser.add_argument('--workbook', default='Upload/full_global_taxonomy_pro (1).xlsx', help='Path to the taxonomy workbook.')
    parser.add_argument('--out', default='data/taxonomy', help='Directory to emit JSON exports.')
    args = parser.parse_args()

    wb_path = Path(args.workbook).expanduser().resolve()
    if not wb_path.exists():
        raise SystemExit(f'Workbook not found: {wb_path}')

    wb = openpyxl.load_workbook(wb_path, data_only=True)
    out_dir = Path(args.out)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        dump_sheet(sheet, out_dir)

    # Print a quick summary for the last tab (assumed to be product items)
    last_sheet = wb[wb.sheetnames[-1]]
    _, records = sheet_to_records(last_sheet)
    summarize_magento_items(records)


if __name__ == '__main__':
    main()
