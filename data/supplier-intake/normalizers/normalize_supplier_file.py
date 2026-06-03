#!/usr/bin/env python3
"""
Supplier-specific file normalizers.

Each normalizer is built from direct inspection of the actual supplier file
structure. Run with --list-profiles to see all available profiles, or
--self-test to validate against sample files.

Usage:
    python normalize_supplier_file.py \\
        --profile ge_tabular_xlsx \\
        --input-file path/to/file.xlsx \\
        --supplier-code GE \\
        > output.csv
"""

from __future__ import annotations

import csv
import re
import sys
import argparse
import tempfile
from pathlib import Path
from typing import Iterator, Any, Optional

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── Canonical output schema ───────────────────────────────────────────────────
# Matches supplier_intake_contract.md and SupplierNormalizedPayload in types.ts
NORMALIZED_FIELDNAMES = [
    "intake_batch_id",
    "source_file_id",
    "source_file_name",
    "source_sheet",
    "source_row_number",
    "supplier_code",
    "supplier_name",
    "supplier_item_code",
    "barcode",
    "product_name",
    "brand",
    "category",
    "country",
    "region",
    "grape",
    "volume_ml",
    "pack_size",
    "vintage",
    "alcohol_pct",
    "cost_ex_vat",
    "cost_inc_vat",
    "supplier_cost",
    "rsp_price",
    "currency",
    "vat_status",
    "discount_pct",
    "raw_price_text",
    "match_key",
    "match_status",
    "matched_sku",
    "proposed_sku",
    "parse_confidence",
    "needs_human_review",
    "validation_errors",
    "notes",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _clean_price(value: Any) -> float | None:
    if value is None or str(value).strip() in ("", "-", "N/A"):
        return None
    try:
        return float(re.sub(r"[^\d.]", "", str(value).replace(",", "")))
    except (ValueError, TypeError):
        return None


def _clean_vol(value: Any) -> int | None:
    """Parse volume to ml. Handles '75cl', '750ml', '0.75', 750."""
    if value is None:
        return None
    s = str(value).strip().lower()
    if not s or s in ("-", "n/a"):
        return None
    try:
        # '75cl' or '37.5cl'
        if "cl" in s:
            return round(float(re.sub(r"[^\d.]", "", s)) * 10)
        # '750ml'
        if "ml" in s:
            return int(float(re.sub(r"[^\d.]", "", s)))
        # fraction like 0.75 (litres)
        f = float(re.sub(r"[^\d.]", "", s))
        if f < 5:
            return round(f * 1000)
        return int(f)
    except (ValueError, TypeError):
        return None


def _clean_vintage(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    # Accept 4-digit year or NV/N.V./nv
    if re.match(r"^(19|20)\d{2}$", s):
        return s
    if s.upper() in ("NV", "N.V.", "NON VINTAGE", "N/V"):
        return "NV"
    # Sometimes Excel gives 2020.0
    m = re.match(r"^(19|20)\d{2}", s)
    if m:
        return m.group(0)
    return None


def _validate(row: dict) -> list[str]:
    errors = []
    if not row.get("product_name"):
        errors.append("missing product_name")
    cost = row.get("supplier_cost") or row.get("cost_ex_vat") or row.get("cost_inc_vat")
    if cost is None:
        errors.append("missing cost")
    elif cost <= 0:
        errors.append("cost_is_zero_or_negative")
    return errors


def _base(supplier_code: str, supplier_name: str, fname: str, sheet: str, row_num: int) -> dict:
    return {
        "intake_batch_id": "",
        "source_file_id": "",
        "source_file_name": fname,
        "source_sheet": sheet,
        "source_row_number": row_num,
        "supplier_code": supplier_code,
        "supplier_name": supplier_name,
        "currency": "THB",
        "match_status": "",
        "matched_sku": "",
        "proposed_sku": "",
    }


def _load_wb(file_path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(file_path, data_only=True, read_only=True)


# ── GE — Great Wine (Thailand) ────────────────────────────────────────────────
# File: List Catalog Greatwine Thailand 3.2026.xlsx
# Sheets: Wine, SPIRITS
# Row 1: header — Code | Type | Names | Grapes | Region | Country | Vintage |
#                 Wholesale Price (THB) | RSP (THB) | Promotion (THB)
# Rows 2+: data. Some rows have no Code (no-code items — include with warning).
# SPIRITS sheet has same structure + Promotion column.

def normalize_ge_tabular_xlsx(file_path: Path, supplier_code: str = "GE") -> Iterator[dict]:
    wb = _load_wb(file_path)
    supplier_name = "Great Wine (Thailand) Company Limited"

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        category_hint = "SPIRITS" if "SPIRIT" in sheet_name.upper() else "WINE"
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row
        header_row_idx = None
        for i, row in enumerate(rows):
            row_s = [str(c).strip() if c else "" for c in row]
            if "Wholesale Price" in " ".join(row_s) or "Wholesale" in " ".join(row_s):
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        header = [str(c).strip() if c else "" for c in rows[header_row_idx]]

        def col(row_vals, *names):
            for name in names:
                for i, h in enumerate(header):
                    if name.lower() in h.lower() and i < len(row_vals):
                        return row_vals[i]
            return None

        for row_num, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            if not any(row):
                continue
            name_val = col(row, "Names", "Name")
            if not name_val:
                continue

            code_val = col(row, "Code")
            cost = _clean_price(col(row, "Wholesale"))
            rsp = _clean_price(col(row, "RSP"))
            errors = []
            if not code_val:
                errors.append("missing_supplier_item_code")
            if cost is None:
                errors.append("missing cost")

            out = _base(supplier_code, supplier_name, file_path.name, sheet_name, row_num)
            out.update({
                "supplier_item_code": str(code_val).strip() if code_val else "",
                "product_name": str(name_val).strip(),
                "category": category_hint,
                "grape": str(col(row, "Grapes", "Grape", "INGREDIENT") or "").strip() or None,
                "region": str(col(row, "Region") or "").strip() or None,
                "country": str(col(row, "Country") or "").strip() or None,
                "vintage": _clean_vintage(col(row, "Vintage")),
                "supplier_cost": cost,
                "rsp_price": rsp,
                "parse_confidence": "high" if code_val and cost else "medium",
                "needs_human_review": not code_val,
                "validation_errors": "|".join(errors) if errors else "",
            })
            yield out


# ── EQ — United Beverage ─────────────────────────────────────────────────────
# File: 4. Quotation May 2026.xlsx
# Sheet: Sheet1 (single sheet)
# Structure: Thai quotation letter. Header rows 3-13 are preamble.
# Row 14: column labels (Thai) — ลำดับ | รูปภาพ | สินค้า | ราคาส่ง ex-VAT | ราคา inc-VAT | RSP
# Row 15: sub-labels (ราคา/ขวด ไม่รวม Vat | ราคาสินค้า รวม VAT 7%)
# Row 16+: data — col A=sequence#, col B=image(skip), col C=product name,
#                 col D=cost_ex_vat, col E=cost_inc_vat, col F=rsp
# Product name is multi-line in cell (name + abv% + size embedded).

def normalize_eq_quotation_xlsx(file_path: Path, supplier_code: str = "EQ") -> Iterator[dict]:
    wb = _load_wb(file_path)
    supplier_name = "United Beverage"
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    # Find first data row: col A is a positive integer (sequence number)
    data_start = None
    for i, row in enumerate(rows):
        if row[0] is not None:
            try:
                n = int(row[0])
                if n >= 1:
                    data_start = i
                    break
            except (TypeError, ValueError):
                pass

    if data_start is None:
        return

    for row_num, row in enumerate(rows[data_start:], start=data_start + 1):
        if not row[0]:
            continue
        try:
            seq = int(row[0])
        except (TypeError, ValueError):
            continue

        name_raw = re.sub(r"\s+", " ", str(row[2])).strip() if row[2] else ""
        if not name_raw:
            continue

        # Parse embedded size from name e.g. "Camus XO 40% / 70cl."
        vol = None
        vol_m = re.search(r"(\d+(?:\.\d+)?)\s*cl", name_raw, re.IGNORECASE)
        if vol_m:
            vol = round(float(vol_m.group(1)) * 10)
        else:
            vol_m2 = re.search(r"(\d+(?:\.\d+)?)\s*ml", name_raw, re.IGNORECASE)
            if vol_m2:
                vol = int(float(vol_m2.group(1)))

        # Parse abv
        abv = None
        abv_m = re.search(r"(\d+(?:\.\d+)?)\s*%", name_raw)
        if abv_m:
            abv = float(abv_m.group(1))

        cost_ex = _clean_price(row[3])
        cost_inc = _clean_price(row[4])
        rsp = _clean_price(row[5])

        errors = _validate({"product_name": name_raw, "cost_ex_vat": cost_ex})

        out = _base(supplier_code, supplier_name, file_path.name, ws.title, row_num)
        out.update({
            "supplier_item_code": str(seq),
            "product_name": name_raw,
            "volume_ml": vol,
            "alcohol_pct": abv,
            "cost_ex_vat": cost_ex,
            "cost_inc_vat": cost_inc,
            "supplier_cost": cost_ex,
            "rsp_price": rsp,
            "vat_status": "inc_vat_available",
            "parse_confidence": "high" if cost_ex and rsp else "medium",
            "needs_human_review": bool(errors),
            "validation_errors": "|".join(errors) if errors else "",
        })
        yield out


# ── AA — Italasia (Wine sheets) ───────────────────────────────────────────────
# File: Wine Price List 20.04.26 update.xlsx
# Sheets: WINES- ACTIVE, EXCLUSIVE WINES, ARTISAN WINES (skip: Price List, ARCHIVE, etc.)
# Structure: repeated header pattern per brand section.
# Section header row: e.g. R1="ITALY", R2="LA KIUVA" (brand name)
# Column header row: Code | (brand) | (region) | Type | Vol. | Vintage | Rating | FB Price | Retail Price | Remark
# Data rows follow until next brand header.
# Key: col A = supplier_item_code (WLK-001R), col B = product name,
#      col C = (region — in header row), col D = type, col E = vol, col F = vintage,
#      col G = rating, col H = FB Price (cost), col I = Retail Price (RSP)
# Supplier codes: AA=wine sheets, AA2=spirits (handled by aa_spirits profile)

DATA_SHEETS_AA = {"WINES- ACTIVE", "EXCLUSIVE WINES", "ARTISAN WINES"}

def normalize_aa_repeated_headers_xlsx(file_path: Path, supplier_code: str = "AA") -> Iterator[dict]:
    wb = _load_wb(file_path)
    supplier_name = "Italasia"

    for sheet_name in wb.sheetnames:
        if sheet_name not in DATA_SHEETS_AA:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        current_country = ""
        current_brand = ""
        current_region = ""
        in_data = False

        for row_num, row in enumerate(rows, start=1):
            # All-empty row resets data block
            non_empty = [c for c in row if c is not None and str(c).strip()]
            if not non_empty:
                in_data = False
                continue

            col_a = str(row[0]).strip() if row[0] else ""
            col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            # Header row: col A == "Code"
            if col_a.lower() == "code":
                # Col C in header row = region label for this brand block
                current_region = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                in_data = True
                continue

            # Country row: col A empty, col B = single UPPERCASE word e.g. "ITALY"
            # (Italasia puts country in col B with col A empty)
            all_other_empty = all(
                row[i] is None or str(row[i]).strip() == ""
                for i in range(2, min(len(row), 9))
            )
            if not col_a and col_b and all_other_empty:
                words = col_b.split()
                if len(words) == 1 and col_b.isupper():
                    current_country = col_b
                    continue
                # Multi-word in col B with col A empty = still a section label (skip silently)
                continue

            if not in_data:
                # Brand name row: col A filled, col B empty, no price cols
                if col_a and not re.match(r"^[A-Z]{2,8}-\d+", col_a):
                    price_cols_empty = all(
                        row[i] is None or str(row[i]).strip() == ""
                        for i in range(7, min(len(row), 9))
                    )
                    if price_cols_empty:
                        current_brand = col_a
                continue

            # Data row: col A = supplier code (or empty for unlisted items)
            name_val = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not name_val:
                # Mid-section brand/region header with no price data
                if col_a and _clean_price(row[7] if len(row) > 7 else None) is None:
                    current_brand = col_a
                    in_data = False
                continue

            cost = _clean_price(row[7]) if len(row) > 7 else None
            rsp  = _clean_price(row[8]) if len(row) > 8 else None
            vol  = _clean_vol(row[4]) if len(row) > 4 else None
            vintage = _clean_vintage(row[5]) if len(row) > 5 else None
            wine_type = str(row[3]).strip() if len(row) > 3 and row[3] else None
            remark = str(row[9]).strip() if len(row) > 9 and row[9] else None

            errors = _validate({"product_name": name_val, "supplier_cost": cost})
            if not col_a:
                errors.append("missing_supplier_item_code")

            out = _base(supplier_code, supplier_name, file_path.name, sheet_name, row_num)
            out.update({
                "supplier_item_code": col_a or "",
                "product_name": name_val,
                "brand": current_brand or None,
                "category": wine_type or "WINE",
                "country": current_country or None,
                "region": current_region or None,
                "volume_ml": vol,
                "vintage": vintage,
                "supplier_cost": cost,
                "rsp_price": rsp,
                "parse_confidence": "high" if col_a and cost else "medium",
                "needs_human_review": not col_a or not cost,
                "validation_errors": "|".join(errors) if errors else "",
                "notes": remark or "",
            })
            yield out


# ── AF — Vanichwattana ────────────────────────────────────────────────────────
# File: WineNow_Fu_Website_January 2026_Final.xlsx
# Sheets: NonWine, Japan, Champ_Spark, Australia, NewZealand, Argentina, Chile,
#         SouthAfrica, USA, Germany, Spain_Hungrian, Italy, France
# Structure per sheet: brand blocks separated by brand name + country row.
# Header row per block: Code | (product name col header) | Vintage | Size | Price/Bottle | Price/NBottles | SRP
# Col F (Price/NBottles) = a negative discount e.g. -0.2 means 20% case discount (ignore for unit pricing)
# Col G = SRP (retail selling price) — sometimes empty
# Data: col A=code (float in Excel e.g. 260001.0), col B=name, col C=vintage,
#       col D=size (e.g. 75cl), col E=Price/Bottle (cost), col G=SRP

SKIP_SHEETS_AF = {"Cover", "PLANOPRAMMING", "CHANNELS"}

def _af_parse_code(v: Any) -> str:
    """Excel stores numeric codes as floats (260001.0). Return as clean int string."""
    if v is None:
        return ""
    try:
        return str(int(float(str(v))))
    except (ValueError, TypeError):
        return str(v).strip()


def normalize_af_multisheet_xlsx(file_path: Path, supplier_code: str = "AF") -> Iterator[dict]:
    """
    Vanichwattana multi-sheet XLSX.

    Most sheets (France, Italy, Champ_Spark, Australia, ...):
      Col A = Code (numeric float), Col B = Name, Col C = Vintage,
      Col D = Size (e.g. 75cl), Col E = Price/Bottle (cost),
      Col F = Price/NBottles (negative discount — skip),
      Col G = SRP (sometimes empty)

    NonWine sheet: offset by 1 — Col B = Code, Col C = Name, Col D = Vintage,
      Col E = Size, Col F = Price/Bottle, Col G = Price/NBottles

    Japan sheet: different structure (Japanese product names, minimal pricing) — emit as needs_human_review.

    Header row detected by: col A (or B for NonWine) == "Code"
    Brand row: row with col B non-empty, no code, no price
    Country row: embedded in brand row as last non-empty cell (e.g. "DEUTZ | France")
    """
    wb = _load_wb(file_path)
    supplier_name = "Vanichwattana"

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS_AF:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # NonWine has code in col B (offset=1); all other sheets code in col A (offset=0)
        is_nonwine = sheet_name == "NonWine"
        is_japan = sheet_name == "Japan"
        offset = 1 if is_nonwine else 0  # code column index

        current_country = ""
        current_brand = ""
        in_data = False

        for row_num, row in enumerate(rows, start=1):
            non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if not non_empty:
                in_data = False
                continue

            code_cell = row[offset] if len(row) > offset else None
            name_cell = row[offset + 1] if len(row) > offset + 1 else None
            code_str = str(code_cell).strip() if code_cell is not None else ""
            name_str = str(name_cell).strip() if name_cell is not None else ""

            # Header row: code column says "Code"
            if code_str.lower() == "code":
                in_data = True
                continue

            # Not yet in a data block — look for brand/country rows
            if not in_data:
                # 1-2 non-empty cells = section header (country or brand+country)
                if len(non_empty) <= 3:
                    # Try to extract brand and country
                    # e.g. row = ['', 'DEUTZ', '', '', '', 'France', ...]
                    text_vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if text_vals:
                        current_brand = text_vals[0]
                        if len(text_vals) > 1:
                            current_country = text_vals[-1]
                continue

            # Japan sheet: emit all rows as needs_human_review
            if is_japan:
                raw_text = " | ".join(str(c) for c in row if c is not None and str(c).strip())
                out = _base(supplier_code, supplier_name, file_path.name, sheet_name, row_num)
                out.update({
                    "product_name": raw_text[:200],
                    "category": "Japan",
                    "parse_confidence": "low",
                    "needs_human_review": True,
                    "validation_errors": "japan_sheet_requires_manual_review",
                })
                yield out
                continue

            # Brand section sub-header (no code, no price — just a brand name row)
            if not code_str or code_str.lower() in ("code", ""):
                if name_str and len(non_empty) <= 3:
                    text_vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    current_brand = text_vals[0]
                    if len(text_vals) > 1:
                        current_country = text_vals[-1]
                    in_data = False
                continue

            if not name_str:
                continue

            # Actual data row
            vintage_cell = row[offset + 2] if len(row) > offset + 2 else None
            size_cell    = row[offset + 3] if len(row) > offset + 3 else None
            cost_cell    = row[offset + 4] if len(row) > offset + 4 else None
            # offset+5 = Price/NBottles (negative discount — skip)
            srp_cell     = row[offset + 6] if len(row) > offset + 6 else None

            cost = _clean_price(cost_cell)
            srp  = _clean_price(srp_cell)

            # Skip the discount rows that look like headers mid-section
            # (e.g. "White Wine", "Red Wine" sub-headers — no code, no price)
            try:
                float(str(code_cell))
            except (TypeError, ValueError):
                # Non-numeric code: could be a sub-category label
                if cost is None:
                    text_vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if text_vals:
                        current_brand = text_vals[0]
                    in_data = False
                    continue

            supplier_item_code = _af_parse_code(code_cell)
            errors = _validate({"product_name": name_str, "supplier_cost": cost})
            if not supplier_item_code:
                errors.append("missing_supplier_item_code")

            out = _base(supplier_code, supplier_name, file_path.name, sheet_name, row_num)
            out.update({
                "supplier_item_code": supplier_item_code,
                "product_name": name_str,
                "brand": current_brand or None,
                "category": sheet_name,
                "country": current_country or None,
                "vintage": _clean_vintage(vintage_cell),
                "volume_ml": _clean_vol(size_cell),
                "supplier_cost": cost,
                "rsp_price": srp,
                "parse_confidence": "high" if supplier_item_code and cost else "medium",
                "needs_human_review": not supplier_item_code or cost is None,
                "validation_errors": "|".join(errors) if errors else "",
            })
            yield out


# ── PDF stubs ─────────────────────────────────────────────────────────────────
# BB&B, IWS, AC (Universal), AH (Ambrose), AE (Gfour), BU (Surawong), FS (SK Liquor)
# are all PDF-only. Automated extraction requires pdfplumber or OCR.
# These stubs emit a single warning row so the run shows up in the review queue.

def _pdf_stub(file_path: Path, supplier_code: str, supplier_name: str) -> Iterator[dict]:
    print(
        f"INFO: {supplier_code} ({supplier_name}) is PDF — "
        "manual extraction required before normalization.",
        file=sys.stderr,
    )
    row = _base(supplier_code, supplier_name, file_path.name, "PDF", 0)
    row.update({
        "product_name": f"[PDF source — {file_path.name}]",
        "parse_confidence": "low",
        "needs_human_review": True,
        "validation_errors": "pdf_source_requires_manual_extraction",
    })
    yield row


def normalize_ab_pdf(file_path: Path, supplier_code: str = "AB") -> Iterator[dict]:
    return _pdf_stub(file_path, supplier_code, "BB&B")

def normalize_ad_pdf(file_path: Path, supplier_code: str = "AD") -> Iterator[dict]:
    return _pdf_stub(file_path, supplier_code, "IWS")

def normalize_ac_pdf(file_path: Path, supplier_code: str = "AC") -> Iterator[dict]:
    return _pdf_stub(file_path, supplier_code, "Universal Fine Wine & Spirit")

def normalize_ah_pdf(file_path: Path, supplier_code: str = "AH") -> Iterator[dict]:
    return _pdf_stub(file_path, supplier_code, "Ambrose")

def normalize_ae_pdf(file_path: Path, supplier_code: str = "AE") -> Iterator[dict]:
    return _pdf_stub(file_path, supplier_code, "Gfour")

def normalize_bu_pdf(file_path: Path, supplier_code: str = "BU") -> Iterator[dict]:
    return _pdf_stub(file_path, supplier_code, "Suriwongse Store")

def normalize_fs_pdf(file_path: Path, supplier_code: str = "FS") -> Iterator[dict]:
    return _pdf_stub(file_path, supplier_code, "SK Liquor")


# ── Registry ──────────────────────────────────────────────────────────────────

NORMALIZERS: dict[str, tuple] = {
    # profile_id: (function, default_supplier_code, description)
    "ge_tabular_xlsx":           (normalize_ge_tabular_xlsx,           "GE",  "Great Wine — tabular XLSX, Code/Wholesale/RSP columns, Wine+SPIRITS sheets"),
    "eq_quotation_xlsx":         (normalize_eq_quotation_xlsx,         "EQ",  "United Beverage — Thai quotation XLSX, cost ex/inc VAT + RSP"),
    "aa_repeated_headers_xlsx":  (normalize_aa_repeated_headers_xlsx,  "AA",  "Italasia Wine — repeated-header XLSX, FB Price/Retail Price, multiple sheets"),
    "af_multisheet_xlsx":        (normalize_af_multisheet_xlsx,        "AF",  "Vanichwattana — multi-sheet XLSX, brand sections, Price/Bottle + SRP"),
    "ab_pdf":                    (normalize_ab_pdf,                    "AB",  "BB&B — PDF source (stub, manual extraction required)"),
    "ad_pdf":                    (normalize_ad_pdf,                    "AD",  "IWS — PDF source (stub, manual extraction required)"),
    "ac_pdf":                    (normalize_ac_pdf,                    "AC",  "Universal Fine Wine — PDF source (stub)"),
    "ah_pdf":                    (normalize_ah_pdf,                    "AH",  "Ambrose — PDF source (stub)"),
    "ae_pdf":                    (normalize_ae_pdf,                    "AE",  "Gfour — PDF source (stub)"),
    "bu_pdf":                    (normalize_bu_pdf,                    "BU",  "Suriwongse Store — PDF source (stub)"),
    "fs_pdf":                    (normalize_fs_pdf,                    "FS",  "SK Liquor — PDF source (stub)"),
}

# Map supplier_code → default profile for convenience
SUPPLIER_DEFAULT_PROFILE: dict[str, str] = {
    "GE":  "ge_tabular_xlsx",
    "EQ":  "eq_quotation_xlsx",
    "AA":  "aa_repeated_headers_xlsx",
    "AA2": "aa_repeated_headers_xlsx",
    "AA4": "aa_repeated_headers_xlsx",
    "AF":  "af_multisheet_xlsx",
    "AB":  "ab_pdf",
    "AB2": "ab_pdf",
    "AB3": "ab_pdf",
    "AD":  "ad_pdf",
    "AC":  "ac_pdf",
    "AH":  "ah_pdf",
    "AE":  "ae_pdf",
    "BU":  "bu_pdf",
    "FS":  "fs_pdf",
}


# ── Self-tests ────────────────────────────────────────────────────────────────

def run_self_tests(sample_dir: Path | None = None) -> bool:
    results = {}

    # GE test
    ge_file = (sample_dir / "GE_May2026.xlsx") if sample_dir else None
    if ge_file and ge_file.exists():
        try:
            rows = list(normalize_ge_tabular_xlsx(ge_file, "GE"))
            assert len(rows) > 20, f"Expected >20 GE rows, got {len(rows)}"
            assert any(r["supplier_item_code"] for r in rows), "No supplier_item_code in GE rows"
            assert any(r["supplier_cost"] for r in rows), "No supplier_cost in GE rows"
            assert any(r["rsp_price"] for r in rows), "No rsp_price in GE rows"
            results["ge_tabular_xlsx"] = f"PASS ({len(rows)} rows)"
        except Exception as e:
            results["ge_tabular_xlsx"] = f"FAIL: {e}"
    else:
        results["ge_tabular_xlsx"] = "SKIP (no sample file)"

    # EQ test
    eq_file = (sample_dir / "EQ_May2026.xlsx") if sample_dir else None
    if eq_file and eq_file.exists():
        try:
            rows = list(normalize_eq_quotation_xlsx(eq_file, "EQ"))
            assert len(rows) > 50, f"Expected >50 EQ rows, got {len(rows)}"
            assert all(r["cost_ex_vat"] for r in rows[:5]), "Missing cost_ex_vat in first EQ rows"
            assert all(r["cost_inc_vat"] for r in rows[:5]), "Missing cost_inc_vat in first EQ rows"
            results["eq_quotation_xlsx"] = f"PASS ({len(rows)} rows)"
        except Exception as e:
            results["eq_quotation_xlsx"] = f"FAIL: {e}"
    else:
        results["eq_quotation_xlsx"] = "SKIP (no sample file)"

    # AA test
    aa_file = (sample_dir / "AA_wine_May2026.xlsx") if sample_dir else None
    if aa_file and aa_file.exists():
        try:
            rows = list(normalize_aa_repeated_headers_xlsx(aa_file, "AA"))
            assert len(rows) > 100, f"Expected >100 AA rows, got {len(rows)}"
            with_code = sum(1 for r in rows if r["supplier_item_code"])
            assert with_code > 50, f"Too few rows with supplier_item_code: {with_code}"
            results["aa_repeated_headers_xlsx"] = f"PASS ({len(rows)} rows, {with_code} with code)"
        except Exception as e:
            results["aa_repeated_headers_xlsx"] = f"FAIL: {e}"
    else:
        results["aa_repeated_headers_xlsx"] = "SKIP (no sample file)"

    # AF test
    af_file = (sample_dir / "AF_vanichwattana.xlsx") if sample_dir else None
    if af_file and af_file.exists():
        try:
            rows = list(normalize_af_multisheet_xlsx(af_file, "AF"))
            assert len(rows) > 200, f"Expected >200 AF rows, got {len(rows)}"
            countries = {r["country"] for r in rows if r["country"]}
            assert len(countries) > 3, f"Expected >3 countries in AF, got {countries}"
            results["af_multisheet_xlsx"] = f"PASS ({len(rows)} rows, {len(countries)} countries)"
        except Exception as e:
            results["af_multisheet_xlsx"] = f"FAIL: {e}"
    else:
        results["af_multisheet_xlsx"] = "SKIP (no sample file)"

    # PDF stubs
    with tempfile.NamedTemporaryFile(suffix=".pdf") as tf:
        pdf_path = Path(tf.name)
        for profile in ["ab_pdf", "ad_pdf", "ac_pdf", "ah_pdf", "ae_pdf", "bu_pdf", "fs_pdf"]:
            try:
                func, code, _ = NORMALIZERS[profile]
                rows = list(func(pdf_path, code))
                assert len(rows) == 1
                assert rows[0]["needs_human_review"]
                results[profile] = "PASS (stub, 1 warning row)"
            except Exception as e:
                results[profile] = f"FAIL: {e}"

    print("\n=== Normalizer Self-Test Results ===", file=sys.stderr)
    all_passed = True
    for name in sorted(results):
        result = results[name]
        status = "✓" if result.startswith("PASS") else ("~" if result.startswith("SKIP") else "✗")
        print(f"  {status} {name:<35} {result}", file=sys.stderr)
        if result.startswith("FAIL"):
            all_passed = False
    print(file=sys.stderr)
    return all_passed


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Normalize supplier price files to canonical CSV schema.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--profile", choices=NORMALIZERS.keys(),
                        help="Normalizer profile to use (see --list-profiles)")
    parser.add_argument("--supplier-code",
                        help="Override supplier code written to output rows")
    parser.add_argument("--input-file", type=Path,
                        help="Path to supplier file (XLSX or PDF)")
    parser.add_argument("--list-profiles", action="store_true",
                        help="Print all available profiles and exit")
    parser.add_argument("--self-test", action="store_true",
                        help="Run self-tests (use --sample-dir to point at XLSX samples)")
    parser.add_argument("--sample-dir", type=Path, default=Path("/tmp/supplier_samples"),
                        help="Directory containing sample XLSX files for self-tests")

    args = parser.parse_args()

    if args.list_profiles:
        print(f"{'Profile':<35} {'Default code':<15} Description")
        print("-" * 90)
        for name, (_, code, desc) in sorted(NORMALIZERS.items()):
            print(f"{name:<35} {code:<15} {desc}")
        sys.exit(0)

    if args.self_test:
        ok = run_self_tests(args.sample_dir if args.sample_dir.exists() else None)
        sys.exit(0 if ok else 1)

    if not args.profile or not args.input_file:
        parser.error("--profile and --input-file are required (or use --self-test / --list-profiles)")

    if not args.input_file.exists():
        print(f"ERROR: File not found: {args.input_file}", file=sys.stderr)
        sys.exit(1)

    func, default_code, _ = NORMALIZERS[args.profile]
    supplier_code = args.supplier_code or default_code

    writer = csv.DictWriter(sys.stdout, fieldnames=NORMALIZED_FIELDNAMES, extrasaction="ignore")
    writer.writeheader()
    count = 0
    for row in func(args.input_file, supplier_code):
        writer.writerow(row)
        count += 1

    print(f"INFO: {count} rows written for {supplier_code} using profile '{args.profile}'",
          file=sys.stderr)
